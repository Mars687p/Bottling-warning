from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from typing import TYPE_CHECKING, Dict
if TYPE_CHECKING:
    from app.reporter.period_worker import TimePeriod


class ExcelWorkbook:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4


    def fill_header(self):
        self.border_header = Border(
                            left=Side(border_style='medium', color='00000000'),
                            right=Side(border_style='medium', color='00000000'),
                            top=Side(border_style='medium', color='00000000'),
                            bottom=Side(border_style='medium', color='00000000'),
                            )
        self.border_body = Border(
                            left=Side(border_style='thin', color='00000000'),
                            right=Side(border_style='thin', color='00000000'),
                            top=Side(border_style='thin', color='00000000'),
                            bottom=Side(border_style='thin', color='00000000'),
                            )
        align = Alignment(vertical='center', horizontal='center', wrap_text=True)
        self.fb10 = Font(name='Times New Roman', size=10, bold=True)
        self.fb11 = Font(name='Times New Roman', size=11, bold=True)
        self.f10 = Font(name='Times New Roman', size=10)

        self.ws['A1'] = 'Линия розлива'
        self.ws['A1'].font = self.fb11
        self.ws['A1'].alignment = align
        self.ws['A1'].border = self.border_header

        self.ws['B1'] = 'Обьем, дал'
        self.ws['B1'].font = self.fb11
        self.ws['B1'].alignment = align
        self.ws['B1'].border = self.border_header

        self.ws['B2'] = 'В раб. время'
        self.ws['B2'].font = self.fb10
        self.ws['B2'].alignment = align
        self.ws['B2'].border = self.border_header

        self.ws['C2'] = 'Сверхурочно'
        self.ws['C2'].font = self.fb10
        self.ws['C2'].alignment = align
        self.ws['C2'].border = self.border_header

        self.ws['D2'] = 'Всего'        
        self.ws['D2'].font = self.fb10
        self.ws['D2'].alignment = align
        self.ws['D2'].border = self.border_header

        self.ws['E1'] = 'Бутылок, шт'
        self.ws['E1'].font = self.fb11
        self.ws['E1'].alignment = align
        self.ws['E1'].border = self.border_header

        self.ws['E2'] = 'В раб. время'
        self.ws['E2'].font = self.fb10
        self.ws['E2'].alignment = align
        self.ws['E2'].border = self.border_header

        self.ws['F2'] = 'Сверхурочно'
        self.ws['F2'].font = self.fb10
        self.ws['F2'].alignment = align
        self.ws['F2'].border = self.border_header

        self.ws['G2'] = 'Всего'
        self.ws['G2'].font = self.fb10
        self.ws['G2'].alignment = align
        self.ws['G2'].border = self.border_header

        self.ws['H1'] = 'Дата'
        self.ws['H1'].font = self.fb11        
        self.ws['H1'].alignment = align
        self.ws['H1'].border = self.border_header
        
        self.ws.merge_cells('A1:A2')
        self.ws.merge_cells('B1:D1')
        self.ws.merge_cells('E1:G1')
        self.ws.merge_cells('H1:H2')

    def fill_footer(self):
        total = self.ws.max_row
        for col in range(1, self.ws.max_column + 1):
            if col == 1:
                cell = self.ws.cell(row=total+1, column=col, value='ИТОГО:')
                cell.font = self.fb11     
                cell.alignment = Alignment(horizontal='right')
                cell.border = self.border_header
                continue

            cell.font = self.fb11     
            cell.alignment = Alignment(horizontal='right')
            cell.border = self.border_header
            if col == self.ws.max_column:
                continue
            cell = self.ws.cell(row=total+1, column=col, 
                    value=f"=Sum({get_column_letter(col)}{3}:{get_column_letter(col)}{total})")

    async def write_wb(self, time_periods: Dict[int, 'TimePeriod']) -> Workbook:
        self.fill_header()

        dim_holder = DimensionHolder(worksheet=self.ws)
        for col in range(self.ws.min_column, self.ws.max_column + 1):
            if col == 1:
                dim_holder[get_column_letter(col)] = ColumnDimension(self.ws, 
                                                                     min=col, max=col, width=25)
            else:
                dim_holder[get_column_letter(col)] = ColumnDimension(self.ws, 
                                                                     min=col, max=col, width=12)
        self.ws.column_dimensions = dim_holder

        for period in time_periods.values():
            for per_date in period.data_per_dates.values():
                for item in per_date.values():
                    await item.calculate_data()
                    self.ws.append([item.line_name, 
                               item.volume_work, item.volume_overtime, item.total_volume, 
                               item.bottles_work, item.bottles_overtime, item.total_bottles,
                               item.date
                               ])
                    
        
        for n_row, row in enumerate(tuple(self.ws.rows), 1):
            if n_row > 2:
                for n_cell, cell in enumerate(row):
                    cell.font = self.f10     
                    # cell.alignment = align
                    cell.border = self.border_body
                    if n_cell == 7:
                        cell.number_format = 'DD.MM.YYYY DDD'
        

        self.fill_footer()
        return self.wb